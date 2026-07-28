import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def export_data_to_excel(sheet_title: str, headers: list, rows: list) -> bytes:
    """
    Generates a styled Excel workbook (.xlsx) from headers and rows.
    Returns bytes buffer.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:30]

    # Header styling
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    # Add Data Rows
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    for r_idx, row_data in enumerate(rows, start=2):
        ws.append(row_data)
        for cell in ws[r_idx]:
            cell.border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
